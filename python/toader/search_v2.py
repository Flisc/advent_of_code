#!/usr/bin/env python3
"""
Search orders, history and stock changes by product name.

File auto-detection:
  - ORDER file         : entries WITHOUT "tip"  → stock OUT (-)
  - HISTORY file       : entries WITH "tip":
      RETUR COMANDA      → stock IN  (+)
      MODIFICARE COMANDA → neutral   (~)
      anything else      → stock OUT (-)
  - STOCK CHANGES file : entries with "stock_operation" field:
      INTRARE_MARFA      → stock IN  (+)
      any other value    → stock OUT (-)
      items searched by items[].label, bucati = bucati_noi

Defaults are set near the top of the file (DEFAULT_* constants).
"""

import argparse
import json
import os
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Default file paths (same directory as script) ───────────────────────────
BASE_DIR =  os.path.dirname(os.path.abspath(__file__))
BASE_EXPORT_DIR = os.path.join(BASE_DIR, "export")

DEFAULT_ORDERS_FILE  = os.path.join(BASE_DIR, "20_03_2026", "tirtoader_orders_2026_MAR-export-20-03.json")
DEFAULT_HISTORY_FILE  = os.path.join(BASE_DIR, "20_03_2026", "tirtoader_deleted_2026_MAR-export-20-03.json")
DEFAULT_STOCK_FILE    = os.path.join(BASE_DIR, "20_03_2026", "stock-changes-MAR-export-20-03.json")

# DEFAULT_ORDERS_FILE = os.path.join(BASE_DIR, "tirtoader_orders_2026_MAR-export.json")
# DEFAULT_HISTORY_FILE = os.path.join(BASE_DIR, "tirtoader_deleted_2026_MAR-export.json")
# DEFAULT_STOCK_FILE = os.path.join(BASE_DIR, "stock_changes-2026-MAR-export.json")
DEFAULT_EXPORT_FILE = os.path.join(BASE_DIR, "export.xlsx")  # set None to disable export

DEFAULT_PRODUCT = "Coarne (10x15x6m)"  # product name to search
# DEFAULT_PRODUCT = "Grinzi (15x15x5m)"  # product name to search
DEFAULT_FROM_DATE = "2026-03-01"  # start date  (YYYY-MM-DD or DD.MM.YYYY)
DEFAULT_SHOW_DELETED = True  # include soft-deleted entries

# ── Stock-direction logic ────────────────────────────────────────────────────
# tip value (lowercase) → (direction_label, sign)
TIP_DIRECTION = {
    "retur comanda": ("RETUR", "+"),  # stock comes back
    "modificare comanda": ("MODIFICARE", "~"),  # neutral edit
    "adaugare_tabel": ("MODIFICARE", "+"),  # neutral edit
    "scadere_tabel": ("MODIFICARE", "-"),  # neutral edit
}
DEFAULT_HISTORY_DIRECTION = ("VANZARE", "-")  # any other tip = sale out

# stock_operation value (lowercase) → (direction_label, sign)
STOCK_OP_DIRECTION = {
    "intrare_marfa": ("INTRARE", "+"),
}
DEFAULT_STOCK_OP_DIRECTION = ("IESIRE", "-")


def stock_op_direction(op):
    """Return (label, sign) for a stock_operation value."""
    return STOCK_OP_DIRECTION.get((op or "").strip().lower(), DEFAULT_STOCK_OP_DIRECTION)


def stock_direction(tip):
    """Return (label, sign) for a given tip value. None means it's a plain order."""
    if tip is None:
        return ("COMANDA", "-")
    return TIP_DIRECTION.get(tip.strip().lower(), DEFAULT_HISTORY_DIRECTION)


# ── Helpers ──────────────────────────────────────────────────────────────────
def get_product_names(item):
    """Extract all product name variants from an order item."""
    names = []
    item_data = item.get("item", {})
    if pn := item_data.get("productName"):
        names.append(pn)
    if label := item_data.get("product", {}).get("label"):
        names.append(label)
    return names


def format_ts(ms):
    try:
        dt = datetime.fromtimestamp(ms / 1000, tz=timezone.utc).astimezone()
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return str(ms)


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# ── Core search ──────────────────────────────────────────────────────────────
def extract_entries(data, search_lower, source_label, show_deleted):
    """
    Parse one file (orders or history) and return matching entries.
    source_label: short string shown in output ("ORDER" / "HISTORY")
    """
    entries = []

    for entry_id, entry in data.items():
        if not show_deleted and entry.get("deletedAt"):
            continue

        matched_products = []
        stock_operation = entry.get("stock_operation")
        tip = ''
        if stock_operation  == "MODIFICARE_TABEL_STOC":
            item = entry.get("changed_item")
            old_item = item.get("old_stock_item")
            new_item = item.get("new_stock_item")
            label = old_item['product']['label'].strip().lower()
            # print("modificare tabel = ", label)
            if label == search_lower:
                print("match tabel = ", label)
                bucati = 0
                if old_item['bucati'] > new_item['bucati']:
                    bucati = old_item['bucati'] - new_item['bucati']
                    tip = 'scadere_tabel'
                if old_item['bucati'] < new_item['bucati']:
                    bucati = new_item['bucati'] - old_item['bucati']
                    tip = 'adaugare_tabel'
                print (tip, bucati)
                direction_label, sign = stock_direction(tip)
                matched_products.append({
                    "productName": label,
                    "label": label,
                    "bucati": bucati,
                    "mc": 0,
                    "pret": 0,
                    "stoc_initial": old_item['bucati'],
                })
        else:
            comanda = entry.get("comanda", {})
            items = comanda.get("items", [])
            tip = entry.get("tip")  # None for plain orders

            direction_label, sign = stock_direction(tip)

            for item in items:
                for name in get_product_names(item):
                    if search_lower in name.lower():
                        label = item.get("item", {}).get("product", {}).get("label", name)
                        matched_products.append({
                            "productName": name,
                            "label": label,
                            "bucati": item.get("bucati"),
                            "mc": item.get("mc"),
                            "pret": item.get("pret"),
                            "stoc_initial": item.get("item").get("bucati"),
                        })
                        break  # one match per item is enough

        if not matched_products:
            continue

        print ('final label ', direction_label, 'sign: ', sign)
        entries.append({
            "entry_id": entry_id,
            "source": source_label,
            "data": entry.get("data") or entry.get("deletedAt", 0),
            "data_str": format_ts(entry.get("data") or entry.get("deletedAt", 0)),
            "clientName": comanda.get("clientName", "—").strip() or "—",
            "tip": tip or "~",
            "direction_label": direction_label,
            "sign": sign,
            "total_comanda": entry.get("total_comanda"),
            "livrat": entry.get("livrat"),
            "incasat": entry.get("incasat"),
            "stoc_initial": entry.get("stoc_initial"),
            "stoc_dupa": entry.get("stoc_dupa_vanzare"),
            "deleted": bool(entry.get("deletedAt")),
            "matched_products": matched_products,
        })

    return entries


def extract_stock_changes(data, search_lower, from_ts=None):
    """
    Parse stock_changes file. Items are flat dicts with:
      label, bucati_noi, nr_buc_initial, volum, ...
    Timestamp field is "date" (same as orders).
    """
    entries = []

    for entry_id, entry in data.items():
        ts = entry.get("date", 0)
        op = entry.get("stock_operation")
        direction_label, sign = stock_op_direction(op)

        matched_products = []
        for item in entry.get("items", []):
            label = item.get("label", "")
            if search_lower in label.lower():
                buc_noi = item.get("bucati_noi", 0)
                buc_initial = item.get("nr_buc_initial", 0)
                volum = item.get("volum", 0.0)
                mc_val = round(buc_noi * volum, 4)
                matched_products.append({
                    "label": label,
                    "bucati": buc_noi,
                    "stoc_initial": buc_initial,
                    "mc": mc_val,
                    "pret": item.get("pret", 0),
                })

        if not matched_products:
            continue

        entries.append({
            "entry_id": entry_id,
            "source": "STOCK",
            "data": ts,
            "data_str": format_ts(ts),
            "clientName": entry.get("furnizor", "—") or "—",
            "tip": op or "—",
            "direction_label": direction_label,
            "sign": sign,
            "total_comanda": None,
            "livrat": None,
            "incasat": None,
            "stoc_initial": entry.get("stoc_initial"),
            "stoc_dupa": entry.get("stoc_final"),
            "deleted": False,
            "matched_products": matched_products,
        })

    return entries


def parse_from_date(date_str):
    """Parse a date string (YYYY-MM-DD or DD.MM.YYYY) into a millisecond timestamp."""
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            dt = datetime.strptime(date_str, fmt).replace(
                hour=0, minute=0, second=0, tzinfo=timezone.utc
            )
            return int(dt.timestamp() * 1000)
        except ValueError:
            continue
    raise argparse.ArgumentTypeError(
        f"Invalid date '{date_str}'. Use YYYY-MM-DD or DD.MM.YYYY"
    )


def search_all(orders_file, history_file, stock_file,
               search_term, show_deleted, from_ts=None):
    results = []
    search_lower = search_term.strip().lower()

    for path, label in [(orders_file, "ORDER"), (history_file, "HISTORY")]:
        if not path:
            continue
        if not os.path.exists(path):
            print(f"  [skip] File not found: {path}")
            continue
        try:
            data = load_json(path)
        except json.JSONDecodeError as e:
            print(f"  [skip] Invalid JSON in {path}: {e}")
            continue
        found = extract_entries(data, search_lower, label, show_deleted)
        results.extend(found)

    if stock_file and os.path.exists(stock_file):
        try:
            stock_data = load_json(stock_file)
            found = extract_stock_changes(stock_data, search_lower)
            results.extend(found)
        except json.JSONDecodeError as e:
            print(f"  [skip] Invalid JSON in {stock_file}: {e}")
    elif stock_file:
        print(f"  [skip] File not found: {stock_file}")

    results.sort(key=lambda x: x["data"])

    if from_ts is not None:
        results = [r for r in results if r["data"] >= from_ts]

    return results


# ── Display ──────────────────────────────────────────────────────────────────
SIGN_COLOR = {"+": "\033[32m", "-": "\033[31m", "~": "\033[33m"}
RESET = "\033[0m"
BOLD = "\033[1m"


def colorize(sign, text):
    return f"{SIGN_COLOR.get(sign, '')}{text}{RESET}"


def print_results(results, search_term, from_date_str=None):
    if not results:
        from_info = f" from {from_date_str}" if from_date_str else ""
        print(f"\n  No entries found for product: '{search_term}'{from_info}\n")
        return

    from_info = f"  from {from_date_str}" if from_date_str else ""
    print(f"\n{BOLD}{'═' * 82}{RESET}")
    print(f"{BOLD}  Chronological stock movements for: '{search_term}'{from_info}  "
          f"({len(results)} entries){RESET}")
    print(f"{BOLD}{'═' * 82}{RESET}\n")

    total_mc_in = 0.0
    total_mc_out = 0.0
    total_buc_in = 0
    total_buc_out = 0

    for i, r in enumerate(results, 1):
        sign = r["sign"]
        deleted = "  ⚠ HISTORIC ACTION" if r["deleted"] else ""
        src_tag = f"[{r['source']}]"

        header = (f"  {colorize(sign, f'[{i}]')}  "
                  f"{BOLD}{r['data_str']}{RESET}  "
                  f"{colorize(sign, sign + ' ' + r['direction_label'])}  "
                  f"{src_tag}{deleted}")
        print(header)
        client_label = "Furnizor" if r["source"] == "STOCK" else "Client "
        print(f"     ID : {r['entry_id']}")
        print(f"       {client_label} : {r['clientName']}")
        print(f"       Tip     : {r['tip']}")

        for p in r["matched_products"]:
            mc_val = p["mc"] or 0.0
            buc_val = p["bucati"] or 0
            if sign == "+":
                total_mc_in += mc_val
                total_buc_in += buc_val
            elif sign == "-":
                total_mc_out += mc_val
                total_buc_out += buc_val

            mc_str = colorize(sign, f"{sign}{mc_val:.3f} mc")
            buc_str = colorize(sign, f"{sign}{buc_val} buc")
            extra = ""
            if r["source"] == "STOCK" and p.get("stoc_initial") is not None:
                extra = f"  | stoc_initial: {p['stoc_initial']} buc"
            else:
                extra = f"  | stoc_initial: {p['stoc_initial']} buc"
            print(f"       Product : {p['label']}"
                  f"  | {buc_str}  | {mc_str} | {extra}")
            # print(f" \t\t {extra}")

        stoc_info = ""
        if r["stoc_initial"] is not None and r["stoc_dupa"] is not None:
            stoc_info = (f"  stoc: {r['stoc_initial']:.3f} → "
                         f"{r['stoc_dupa']:.3f} mc")
        if r["total_comanda"] is not None:
            print(f"       Total   : {r['total_comanda']} RON{stoc_info}")
        elif stoc_info:
            print(f"       Stoc    :{stoc_info}")
        print()

    # Summary
    net_mc = total_mc_in - total_mc_out
    net_buc = total_buc_in - total_buc_out
    net_mc_str = colorize("+" if net_mc >= 0 else "-", f"{net_mc:+.3f} mc")
    net_buc_str = colorize("+" if net_buc >= 0 else "-", f"{net_buc:+d} buc")
    print(f"{'-' * 82}")
    print(f"  Summary   IN : {colorize('+', f'+{total_mc_in:.3f} mc')}  {colorize('+', f'+{total_buc_in} buc')}   "
          f"OUT: {colorize('-', f'-{total_mc_out:.3f} mc')}  {colorize('-', f'-{total_buc_out} buc')}   "
          f"NET: {net_mc_str}  {net_buc_str}")
    print(f"{'═' * 82}\n")


# ── Excel export ─────────────────────────────────────────────────────────────
SIGN_LABEL = {"+": "IN", "-": "OUT", "~": "NEUTRAL"}

# Excel colors
COLOR_IN = "C6EFCE"  # green fill
COLOR_OUT = "FFC7CE"  # red fill
COLOR_NEUTRAL = "FFEB9C"  # yellow fill
COLOR_HEADER = "2E4057"  # dark blue header bg
COLOR_STOCK = "DDEEFF"  # light blue for STOCK rows
COLOR_SUMMARY = "F2F2F2"  # light grey summary


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value, bold=False, fill_hex=None, number_format=None,
          align="left", font_color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=font_color, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    c.border = _thin_border()
    if fill_hex:
        c.fill = PatternFill("solid", start_color=fill_hex, end_color=fill_hex)
    if number_format:
        c.number_format = number_format
    return c


def export_to_excel(results, search_term, from_date_str, export_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Movements"
    ws.freeze_panes = "A2"

    # ── Headers ───────────────────────────────────────────────────────────────
    headers = ["#", "Date", "Source", "Direction", "Client / Furnizor",
               "Tip", "Product", "Bucati", "MC", "Stoc Initial", "Stoc Final",
               "Total RON", "Buc Initial (stoc)"]
    for col, h in enumerate(headers, 1):
        c = _cell(ws, 1, col, h, bold=True, fill_hex=COLOR_HEADER,
                  align="center", font_color="FFFFFF")

    col_widths = [4, 20, 8, 12, 22, 22, 28, 8, 10, 12, 10, 12, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Data rows ─────────────────────────────────────────────────────────────
    row = 2
    total_mc_in = total_mc_out = 0.0
    total_buc_in = total_buc_out = 0

    sign_fill = {"+": COLOR_IN, "-": COLOR_OUT, "~": COLOR_NEUTRAL}

    data_start_row = row  # for summary formulas later

    for i, r in enumerate(results, 1):
        sign = r["sign"]
        fill = COLOR_STOCK if r["source"] == "STOCK" else sign_fill.get(sign, "FFFFFF")
        products = r["matched_products"]
        n = len(products)

        for pi, p in enumerate(products):
            mc_val = p["mc"] or 0.0
            buc_val = p["bucati"] or 0

            if sign == "+":
                total_mc_in += mc_val;
                total_buc_in += buc_val
            elif sign == "-":
                total_mc_out += mc_val;
                total_buc_out += buc_val

            # Only fill left columns once per entry (first product row)
            is_first = pi == 0
            _cell(ws, row, 1, i if is_first else "", fill_hex=fill, align="center")
            _cell(ws, row, 2, r["data_str"] if is_first else "", fill_hex=fill)
            _cell(ws, row, 3, r["source"] if is_first else "", fill_hex=fill, align="center")
            _cell(ws, row, 4, SIGN_LABEL.get(sign, sign) if is_first else "", fill_hex=fill, align="center")
            _cell(ws, row, 5, r["clientName"] if is_first else "", fill_hex=fill)
            _cell(ws, row, 6, r["tip"] if is_first else "", fill_hex=fill)
            _cell(ws, row, 7, p["label"], fill_hex=fill)
            _cell(ws, row, 8, buc_val, fill_hex=fill, align="right", number_format="#,##0")
            _cell(ws, row, 9, mc_val, fill_hex=fill, align="right", number_format="#,##0.000")
            _cell(ws, row, 10, r["stoc_initial"] if is_first and r.get("stoc_initial") else
            (r["stoc_initial"] if is_first else ""),
                  fill_hex=fill, align="right", number_format="#,##0.000")
            _cell(ws, row, 11, r["stoc_dupa"] if is_first else "", fill_hex=fill, align="right",
                  number_format="#,##0.000")
            _cell(ws, row, 12, r["total_comanda"] if is_first else "", fill_hex=fill, align="right",
                  number_format="#,##0.00")
            _cell(ws, row, 13, p.get("stoc_initial", ""), fill_hex=fill, align="right", number_format="#,##0")
            row += 1

    # ── Summary row ───────────────────────────────────────────────────────────
    row += 1  # blank spacer
    sum_fill = COLOR_SUMMARY
    _cell(ws, row, 1, "TOTAL", bold=True, fill_hex=sum_fill)
    _cell(ws, row, 2, f"Entries: {len(results)}", fill_hex=sum_fill)
    _cell(ws, row, 3, f"IN mc: +{total_mc_in:.3f}", bold=True, fill_hex=COLOR_IN)
    _cell(ws, row, 4, f"IN buc: +{total_buc_in}", bold=True, fill_hex=COLOR_IN)
    _cell(ws, row, 5, f"OUT mc: -{total_mc_out:.3f}", bold=True, fill_hex=COLOR_OUT)
    _cell(ws, row, 6, f"OUT buc: -{total_buc_out}", bold=True, fill_hex=COLOR_OUT)
    net_mc = total_mc_in - total_mc_out
    net_buc = total_buc_in - total_buc_out
    net_fill = COLOR_IN if net_mc >= 0 else COLOR_OUT
    _cell(ws, row, 7, f"NET mc: {net_mc:+.3f}", bold=True, fill_hex=net_fill)
    _cell(ws, row, 8, f"NET buc: {net_buc:+d}", bold=True, fill_hex=net_fill)

    # ── Raw data sheet ───────────────────────────────────────────────────────
    rs = wb.create_sheet("Raw Data")

    # Collect all unique product-level keys across all results
    product_keys = ["label", "productName", "bucati", "mc", "pret", "buc_initial"]

    raw_headers = [
        "entry_id", "source", "data_str", "direction_label", "sign",
        "clientName", "tip", "total_comanda", "livrat", "incasat",
        "stoc_initial", "stoc_dupa", "deleted",
        # product fields (flattened)
        "product_label", "product_productName", "product_bucati",
        "product_mc", "product_pret", "product_buc_initial",
    ]

    for col, h in enumerate(raw_headers, 1):
        c = rs.cell(row=1, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", start_color=COLOR_HEADER, end_color=COLOR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin_border()

    rs.freeze_panes = "A2"

    raw_row = 2
    for r in results:
        sign = r["sign"]
        fill_hex = sign_fill.get(sign, "FFFFFF")
        if r["source"] == "STOCK":
            fill_hex = COLOR_STOCK

        products = r["matched_products"]
        for pi, p in enumerate(products):
            is_first = pi == 0
            vals = [
                r["entry_id"] if is_first else "",
                r["source"] if is_first else "",
                r["data_str"] if is_first else "",
                r["direction_label"] if is_first else "",
                r["sign"] if is_first else "",
                r["clientName"] if is_first else "",
                r["tip"] if is_first else "",
                r["total_comanda"] if is_first else "",
                r["livrat"] if is_first else "",
                r["incasat"] if is_first else "",
                r["stoc_initial"] if is_first else "",
                r["stoc_dupa"] if is_first else "",
                r["deleted"] if is_first else "",
                # product fields
                p.get("label", ""),
                p.get("productName", ""),
                p.get("bucati", ""),
                p.get("mc", ""),
                p.get("pret", ""),
                p.get("stoc_initial", ""),
            ]
            for col, val in enumerate(vals, 1):
                c = rs.cell(row=raw_row, column=col, value=val)
                c.font = Font(name="Arial", size=10)
                c.fill = PatternFill("solid", start_color=fill_hex, end_color=fill_hex)
                c.border = _thin_border()
                c.alignment = Alignment(vertical="center")
            raw_row += 1

    # auto-width for raw sheet
    raw_col_widths = [28, 8, 20, 12, 6, 22, 22, 12, 8, 8, 12, 12, 8, 28, 16, 8, 10, 8, 18]
    for i, w in enumerate(raw_col_widths, 1):
        rs.column_dimensions[get_column_letter(i)].width = w

    # ── Meta sheet ────────────────────────────────────────────────────────────
    ms = wb.create_sheet("Info")
    ms["A1"] = "Search term";
    ms["B1"] = search_term
    ms["A2"] = "From date";
    ms["B2"] = from_date_str or "—"
    ms["A3"] = "Exported at";
    ms["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ms["A4"] = "Total entries";
    ms["B4"] = len(results)
    for cell in ["A1", "A2", "A3", "A4"]:
        ms[cell].font = Font(bold=True)

    wb.save(export_path)
    return export_path

# ── Entry point ──────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Search orders & history by product, sorted chronologically.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Stock direction:
  ORDER file (no 'tip')          →  - OUT  (sale / delivery)
  HISTORY tip=RETUR COMANDA      →  + IN   (stock return)
  HISTORY tip=MODIFICARE COMANDA →  ~ NEUTRAL (edit, no stock change)
  HISTORY any other tip          →  - OUT

Examples:
  python search_orders.py Coarne
  python search_orders.py "Mini Coarne" --show-deleted
  python search_orders.py Leturi --orders-file data/orders.json --history-file data/hist.json
        """
    )
    parser.add_argument("product", nargs="?", default=DEFAULT_PRODUCT,
                        help=f"Product name to search, partial/case-insensitive "
                             f"(default: '{DEFAULT_PRODUCT}')")
    parser.add_argument("--show-deleted", action="store_true",
                        default=DEFAULT_SHOW_DELETED,
                        help=f"Include soft-deleted entries "
                             f"(default: {DEFAULT_SHOW_DELETED})")
    parser.add_argument("--from-date", default=DEFAULT_FROM_DATE,
                        type=parse_from_date, metavar="DATE",
                        help=f"Only show entries on or after this date. "
                             f"Formats: YYYY-MM-DD or DD.MM.YYYY "
                             f"(default: {DEFAULT_FROM_DATE})")
    parser.add_argument("--orders-file", default=DEFAULT_ORDERS_FILE,
                        help=f"Path to orders JSON        (default: {DEFAULT_ORDERS_FILE})")
    parser.add_argument("--history-file", default=DEFAULT_HISTORY_FILE,
                        help=f"Path to history JSON       (default: {DEFAULT_HISTORY_FILE})")
    parser.add_argument("--stock-file", default=DEFAULT_STOCK_FILE,
                        help=f"Path to stock_changes JSON (default: {DEFAULT_STOCK_FILE})")
    parser.add_argument("--export-file", default=DEFAULT_EXPORT_FILE,
                        help=f"Path for Excel export; set empty string to disable "
                             f"(default: {DEFAULT_EXPORT_FILE})")

    args = parser.parse_args()

    results = search_all(
        orders_file=args.orders_file,
        history_file=args.history_file,
        stock_file=args.stock_file,
        search_term=args.product,
        show_deleted=args.show_deleted,
        from_ts=args.from_date,
    )
    from_date_str = args.from_date and datetime.fromtimestamp(
        args.from_date / 1000, tz=timezone.utc
    ).strftime("%Y-%m-%d")
    print_results(results, args.product, from_date_str=from_date_str)

    if args.export_file and results:
        now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        path = f"{args.product}-istoric-stoc-MAR-{now}.xlsx"
        export_path = os.path.join(BASE_EXPORT_DIR, path)
        os.makedirs(BASE_EXPORT_DIR, exist_ok=True)
        path = export_to_excel(results, args.product, from_date_str, export_path)
        print(f"  Excel exported → {export_path}")
    elif args.export_file and not results:
        print("  No results to export.")

if __name__ == "__main__":
    main()
